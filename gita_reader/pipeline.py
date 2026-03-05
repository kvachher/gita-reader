#!/usr/bin/env python3
"""Generate per-person to-do lists from the MM19 planning workbook."""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable

import openpyxl

GROUP_BOARD = "__ALL_BOARD__"
GROUP_LIAISON = "__ALL_LIAISONS__"
GROUP_VOLUNTEER = "__ALL_VOLUNTEERS__"

LOGISTICS_SHEETS = [
    "Wednesday Logistics",
    "Thursday Logistics",
    "Friday Logistics",
    "Saturday Logistics",
    "Sunday Logistics",
]
DAY_ORDER = {name: idx for idx, name in enumerate(LOGISTICS_SHEETS)}

NICKNAME_ALIASES = {
    "param": "Param Jhala",
    "vaishu": "Vaishnavi Chintagumpala",
    "singhvi": "Aditya Singhvi",
    "kadaru": "Riya Kadaru",
    "sista": "Sriman Sista",
    "banda": "Vaishnavi Banda",
    "gunna": "Navya Gunna",
    "megha": "Meghana Puri",
    "sisterkavya": "Kavya Shah",
    "prajapati": "Param Prajapati",
}

TRAVEL_SHEETS = {
    "Friday Travel Logistics",
    "Saturday Travel Logistics",
    "Sunday Travel Logistics",
}

ASSIGNMENT_SHEETS = {
    "Friday Board Assignments (Mixer",
    "Saturday Board Assignments",
}

OTHER_ROLE_SHEETS = {
    "Show Schedule",
    "Post-Show Feedback",
}

SECTION_BOARD = "board"
SECTION_TEAM_LIAISONS = "team_liaisons"


def norm(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%I:%M %p").lstrip("0")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


@dataclass
class Task:
    sheet: str
    when: str
    summary: str
    details: str
    role: str
    parsed: dict[str, str] | None = None

    def to_dict(self) -> dict[str, object]:
        payload: dict[str, object] = {
            "sheet": self.sheet,
            "when": self.when,
            "summary": self.summary,
            "details": self.details,
            "role": self.role,
        }
        if self.parsed:
            payload["parsed"] = self.parsed
        return payload


class TodoBuilder:
    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        self.wb = openpyxl.load_workbook(workbook_path, data_only=True)
        self.people: set[str] = set()
        self.alias_to_name: dict[str, str] = {}
        self.board_members: set[str] = set()
        self.liaisons: set[str] = set()
        self.volunteers: set[str] = set()
        self.person_category: dict[str, str] = {}
        self.person_committee: dict[str, str] = {}
        self.important_info: list[dict[str, str]] = []
        self.contacts: list[dict[str, str]] = []
        self.tasks_by_assignee: dict[str, list[Task]] = defaultdict(list)
        self.calendar_by_sheet: dict[str, dict[str, object]] = {}
        self.nickname_aliases: dict[str, str] = {}

    def build(self) -> dict[str, object]:
        self._extract_contacts()
        self._extract_important_info()
        self._extract_logistics_tasks()
        self._extract_calendar()
        self._extract_assignments()
        self._extract_travel()
        self._extract_other_roles()
        self._propagate_group_tasks()
        return self._to_export()

    def _extract_contacts(self) -> None:
        ws = self.wb["Board and Liaison Contacts"]
        section = ""
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=7, values_only=True):
            first_col = norm(row[0])
            name_col_2 = norm(row[1])
            cap_or_name_col_3 = norm(row[2])
            role_col_4 = norm(row[3])
            liaison_col_5 = norm(row[4])
            liaison_phone_col_6 = norm(row[5]) if len(row) > 5 else ""
            liaison_phone_col_7 = norm(row[6]) if len(row) > 6 else ""

            marker = first_col.upper()
            if "EXEC BOARD" in marker:
                section = SECTION_BOARD
                continue
            if "TEAMS + CAPTAINS + LIAISONS" in marker:
                section = SECTION_TEAM_LIAISONS
                continue
            if marker in {"RAS + JUDGES", "DJ + MEDIA", "VOLUNTEERS"}:
                section = marker
                continue

            if section == SECTION_BOARD:
                self._register_person(
                    name_col_2,
                    category="board",
                    committee=role_col_4,
                    role=role_col_4,
                    phone=liaison_col_5,
                    section="Board",
                    board=True,
                    liaison=("liaison" in role_col_4.lower()),
                )
                continue
            if section == SECTION_TEAM_LIAISONS:
                self._register_person(
                    cap_or_name_col_3,
                    category="other",
                    role="Team Captain",
                    phone=role_col_4,
                    section="Teams",
                )
                self._register_person(
                    liaison_col_5,
                    category="liaison",
                    role="Liaison",
                    phone=liaison_phone_col_7 or liaison_phone_col_6,
                    section="Teams",
                    liaison=True,
                )
                continue
            if section in {"RAS + JUDGES", "DJ + MEDIA", "VOLUNTEERS"}:
                if section == "VOLUNTEERS":
                    self._register_person(
                        name_col_2,
                        category="volunteer",
                        role=liaison_col_5,
                        phone=role_col_4,
                        section="Volunteers",
                        volunteer=True,
                    )
                else:
                    self._register_person(
                        name_col_2,
                        category="other",
                        role=liaison_col_5,
                        phone=role_col_4,
                        section=section.title(),
                    )
                continue

        self.alias_to_name = {k: v for k, v in self.alias_to_name.items() if v}
        for alias, canonical in NICKNAME_ALIASES.items():
            if canonical in self.people:
                self.alias_to_name[alias.lower()] = canonical
                self.nickname_aliases[alias.lower()] = canonical

    def _extract_important_info(self) -> None:
        ws = self.wb["Important Information"]
        header = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=4, values_only=True):
            vals = [norm(v) for v in row]
            if not any(vals):
                continue
            if vals[0].lower() == "item" and vals[1].lower() == "location":
                header = "locations"
                continue
            if vals[0].lower() in {"item", "location", "address"}:
                continue
            if header == "locations" and vals[0]:
                if vals[0].lower() in {"name", "all applications"}:
                    continue
                self.important_info.append(
                    {
                        "item": vals[0],
                        "location": vals[1],
                        "address": vals[2],
                    }
                )

    def _extract_logistics_tasks(self) -> None:
        for sheet_name in LOGISTICS_SHEETS:
            ws = self.wb[sheet_name]
            header_row = None
            headers = {}

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
                values = [norm(v) for v in row]
                if "TIME" in values:
                    header_row = row_idx
                    headers = {i: values[i].strip().upper() for i in range(len(values)) if values[i]}
                    continue
                if header_row is None or row_idx <= header_row:
                    continue

                row_has_content = any(values)
                if not row_has_content:
                    continue

                when = values[0]
                if not when:
                    when = self._find_nearest_time(ws, row_idx)

                for col_idx, value in enumerate(values):
                    if not value:
                        continue
                    col_header = headers.get(col_idx, "")
                    role = col_header or "General"
                    self._assign_from_text(
                        sheet=sheet_name,
                        when=when,
                        role=role,
                        text=value,
                        row_values=values,
                    )

    def _extract_calendar(self) -> None:
        for sheet_name in LOGISTICS_SHEETS:
            ws = self.wb[sheet_name]
            header_row = None
            headers: list[str] = []
            day_title = ""
            events: list[dict[str, object]] = []

            for r in range(1, 8):
                line = norm(ws.cell(r, 2).value)
                if line and "logistics" in line.lower():
                    day_title = line
                    break

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
                values = [norm(v) for v in row]
                if "TIME" in values:
                    header_row = row_idx
                    headers = [v for v in values if v]
                    continue
                if header_row is None or row_idx <= header_row:
                    continue

                time_value = values[0] if values else ""
                row_cells = []
                for col_idx in range(1, len(values)):
                    cell_text = values[col_idx]
                    if not cell_text:
                        continue
                    col_name = ""
                    if col_idx < len(row) and ws.cell(header_row, col_idx + 1).value is not None:
                        col_name = norm(ws.cell(header_row, col_idx + 1).value)
                    row_cells.append({"column": col_name or f"Column {col_idx+1}", "text": cell_text})

                if row_cells:
                    events.append({"time": time_value, "cells": row_cells})

            self.calendar_by_sheet[sheet_name] = {
                "sheet": sheet_name,
                "title": day_title or sheet_name,
                "headers": headers,
                "events": events,
            }

    def _extract_assignments(self) -> None:
        for sheet_name in ASSIGNMENT_SHEETS:
            ws = self.wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6, values_only=True):
                role = norm(row[0])
                name_cell = norm(row[1])
                classification = norm(row[2])
                details = norm(row[3])

                if not role or role.lower().startswith("roles"):
                    continue
                if not name_cell or name_cell.lower() == "name":
                    continue

                person = self._resolve_name(name_cell)
                if person:
                    self._add_task(
                        person,
                        Task(
                            sheet=sheet_name,
                            when="",
                            summary=role,
                            details=details,
                            role=classification or "Assignment",
                        ),
                    )

    def _extract_travel(self) -> None:
        for sheet_name in TRAVEL_SHEETS:
            ws = self.wb[sheet_name]
            header_map = {}
            for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
                values = [norm(v) for v in row]
                lowered = [v.lower() for v in values]
                if "driver" in " ".join(lowered):
                    header_map = {values[i].lower(): i for i in range(len(values)) if values[i]}
                    break

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=12, values_only=True):
                values = [norm(v) for v in row]
                if not any(values):
                    continue

                if sheet_name == "Friday Travel Logistics":
                    driver_idx = header_map.get("driver", 10)
                    time_idx = header_map.get("time", 1)
                    team_idx = header_map.get("team", 4)
                    first_idx = header_map.get("name (first)", 2)
                    last_idx = header_map.get("name (last)", 3)
                    driver = values[driver_idx] if len(values) > driver_idx else ""
                    pickup_time = values[time_idx] if len(values) > time_idx else ""
                    team = values[team_idx] if len(values) > team_idx else ""
                    rider_name = " ".join(v for v in [values[first_idx], values[last_idx]] if v)
                    if driver and driver.lower() not in {"driver", "driver name"}:
                        person = self._resolve_name(driver)
                        if person:
                            self._add_task(
                                person,
                                Task(
                                    sheet=sheet_name,
                                    when=pickup_time,
                                    summary=f"Airport pickup for {team}".strip(),
                                    details=f"Rider: {rider_name}".strip(),
                                    role="Driver",
                                ),
                            )
                else:
                    driver_idx = header_map.get("driver name", 4)
                    depart_idx = header_map.get("departure time", header_map.get("time to leave for the airport", 1))
                    route_from_idx = header_map.get("from location", 7)
                    route_to_idx = header_map.get("to location", 8)
                    team_idx = header_map.get("team to transport", header_map.get("team name", 10))
                    driver = values[driver_idx] if len(values) > driver_idx else ""
                    when = values[depart_idx] if len(values) > depart_idx else ""
                    route_from = values[route_from_idx] if len(values) > route_from_idx else ""
                    route_to = values[route_to_idx] if len(values) > route_to_idx else ""
                    team = values[team_idx] if len(values) > team_idx else ""
                    if driver and driver.lower() not in {"driver name", ""}:
                        person = self._resolve_name(driver)
                        if person:
                            self._add_task(
                                person,
                                Task(
                                    sheet=sheet_name,
                                    when=when,
                                    summary=f"Transport: {team}".strip(),
                                    details=f"Route: {route_from} -> {route_to}".strip(),
                                    role="Driver",
                                ),
                            )

    def _extract_other_roles(self) -> None:
        # Show schedule POC assignments
        show_ws = self.wb["Show Schedule"]
        for row in show_ws.iter_rows(min_row=1, max_row=show_ws.max_row, max_col=9, values_only=True):
            values = [norm(v) for v in row]
            if len(values) < 8:
                continue
            start = values[1]
            activity = values[3]
            poc = values[7]
            if poc and poc.lower() != "poc":
                person = self._resolve_name(poc)
                if person:
                    self._add_task(
                        person,
                        Task(
                            sheet="Show Schedule",
                            when=start,
                            summary=activity or "Show task",
                            details="POC assignment",
                            role="POC",
                        ),
                    )

        # Post-show logistics contact in row text
        post_ws = self.wb["Post-Show Feedback"]
        for row in post_ws.iter_rows(min_row=1, max_row=post_ws.max_row, max_col=10, values_only=True):
            values = [norm(v) for v in row]
            if len(values) < 8:
                continue
            when = values[2]
            contact = values[7]
            if contact and "@" not in contact and "contact" not in contact.lower():
                person = self._resolve_name(contact)
                if person:
                    self._add_task(
                        person,
                        Task(
                            sheet="Post-Show Feedback",
                            when=when,
                            summary="Post-show logistics contact",
                            details=f"Team slot: {values[1]}",
                            role="Logistics",
                        ),
                    )

    def _propagate_group_tasks(self) -> None:
        board_tasks = self.tasks_by_assignee.get(GROUP_BOARD, [])
        liaison_tasks = self.tasks_by_assignee.get(GROUP_LIAISON, [])
        volunteer_tasks = self.tasks_by_assignee.get(GROUP_VOLUNTEER, [])
        for person in self.board_members:
            self.tasks_by_assignee[person].extend(board_tasks)
        for person in self.liaisons:
            self.tasks_by_assignee[person].extend(liaison_tasks)
        for person in self.volunteers:
            self.tasks_by_assignee[person].extend(volunteer_tasks)

    def _assign_from_text(self, sheet: str, when: str, role: str, text: str, row_values: list[str]) -> None:
        if not text:
            return
        segments = self._split_cell_segments(text, when)
        for segment in segments:
            segment_text = segment["text"]
            parsed = self._parse_task_text(segment_text)
            parsed["full_text"] = text
            effective_when = segment.get("when") or parsed.get("time") or when

            if self._is_group_board_task(role, segment_text):
                self._add_task(
                    GROUP_BOARD,
                    Task(
                        sheet=sheet,
                        when=effective_when,
                        summary=segment_text,
                        details="Shared Board responsibility",
                        role=role,
                        parsed=parsed,
                    ),
                )
            if self._is_group_liaison_task(role, segment_text):
                self._add_task(
                    GROUP_LIAISON,
                    Task(
                        sheet=sheet,
                        when=effective_when,
                        summary=segment_text,
                        details="Shared Liaison responsibility",
                        role=role,
                        parsed=parsed,
                    ),
                )
            if self._is_group_volunteer_task(role, segment_text):
                self._add_task(
                    GROUP_VOLUNTEER,
                    Task(
                        sheet=sheet,
                        when=effective_when,
                        summary=segment_text,
                        details="Shared Volunteer responsibility",
                        role=role,
                        parsed=parsed,
                    ),
                )

            explicit = self._extract_named_people(segment_text)
            if explicit:
                for person in explicit:
                    self._add_task(
                        person,
                        Task(
                            sheet=sheet,
                            when=effective_when,
                            summary=segment_text,
                            details="",
                            role=role,
                            parsed=parsed,
                        ),
                    )

    def _split_cell_segments(self, text: str, fallback_when: str) -> list[dict[str, str]]:
        cleaned = text.replace("\r\n", "\n").strip()
        matches = list(re.finditer(r"\[(?P<time>[^\]]+)\]", cleaned))
        if not matches:
            return [{"when": fallback_when, "text": cleaned}]

        segments: list[dict[str, str]] = []
        prefix = cleaned[: matches[0].start()].strip()
        if prefix:
            segments.append({"when": fallback_when, "text": prefix})

        for idx, match in enumerate(matches):
            start = match.start()
            end = matches[idx + 1].start() if idx + 1 < len(matches) else len(cleaned)
            chunk = cleaned[start:end].strip()
            if not chunk:
                continue
            segments.append({"when": match.group("time").strip(), "text": chunk})
        return segments

    def _parse_task_text(self, text: str) -> dict[str, str]:
        normalized = text.replace("\r\n", "\n").strip()
        lines = [line.strip() for line in normalized.split("\n") if line.strip()]
        if not lines:
            return {"format": "raw", "raw": text}

        first = lines[0]
        m = re.match(r"^\[(?P<time>[^\]]+)\]\s*(?P<title>.+)$", first)
        if not m:
            return {"format": "raw", "raw": text, "full_text": text}

        parsed: dict[str, str] = {
            "format": "titled",
            "time": m.group("time").strip(),
            "title": m.group("title").strip(),
            "full_text": text,
        }
        body_lines = lines[1:]
        while body_lines and body_lines[0] == first:
            body_lines = body_lines[1:]
        if body_lines:
            dup_head = re.match(r"^\[(?P<time>[^\]]+)\]\s*(?P<title>.+)$", body_lines[0])
            if dup_head and dup_head.group("time").strip() == parsed["time"] and dup_head.group("title").strip() == parsed["title"]:
                body_lines = body_lines[1:]
        body = "\n".join(body_lines).strip()
        parsed["body"] = body
        if not body:
            parsed["format"] = "raw"
            parsed["raw"] = text
            return parsed

        # (1) departure format: People in Car to Place from Place (for Thing)
        dep = re.search(
            r"(?P<people>.+?)\s+in\s+(?P<car>.+?)\s+to\s+(?P<to>.+?)\s+from\s+(?P<from>.+?)(?:\s+for\s+(?P<purpose>.+))?$",
            body,
            re.IGNORECASE,
        )
        if dep:
            parsed["format"] = "route_depart"
            parsed["people"] = dep.group("people").strip()
            parsed["car"] = dep.group("car").strip()
            parsed["to"] = dep.group("to").strip()
            parsed["from"] = dep.group("from").strip()
            if dep.group("purpose"):
                parsed["purpose"] = dep.group("purpose").strip()
            return parsed

        # (2) arrival format: People in Car arrive(s) at Place (for Thing)
        arr = re.search(
            r"(?P<people>.+?)\s+in\s+(?P<car>.+?)\s+arrive[s]?\s+at\s+(?P<at>.+?)(?:\s+for\s+(?P<purpose>.+))?$",
            body,
            re.IGNORECASE,
        )
        if arr:
            parsed["format"] = "route_arrive"
            parsed["people"] = arr.group("people").strip()
            parsed["car"] = arr.group("car").strip()
            parsed["at"] = arr.group("at").strip()
            if arr.group("purpose"):
                parsed["purpose"] = arr.group("purpose").strip()
            return parsed

        # (3) fallback raw body
        parsed["format"] = "raw"
        parsed["raw"] = text
        return parsed

    def _find_nearest_time(self, ws, row_idx: int) -> str:
        for i in range(row_idx, max(1, row_idx - 6), -1):
            val = norm(ws.cell(i, 1).value)
            if val:
                return val
        return ""

    def _is_group_board_task(self, role: str, text: str) -> bool:
        test = f"{role} {text}".upper()
        return "ALL_BOARD" in test or "ALL BOARD" in test

    def _is_group_liaison_task(self, role: str, text: str) -> bool:
        test = f"{role} {text}".upper()
        return "ALL_LIAISON" in test or "ALL LIAISON" in test or "ALL LIAISONS" in test

    def _is_group_volunteer_task(self, role: str, text: str) -> bool:
        test = f"{role} {text}".upper()
        return "ALL_VOLUNTEER" in test or "ALL VOLUNTEER" in test or "ALL VOLUNTEERS" in test

    def _extract_named_people(self, text: str) -> list[str]:
        cleaned = re.sub(r"[\[\](){}*]", " ", text)
        tokens = re.split(r"[,/+]| and | & |\n|;", cleaned)
        matches = []
        for token in tokens:
            token = token.strip()
            if not token:
                continue
            person = self._resolve_name(token)
            if person:
                matches.append(person)

        # Phrase-level contains check for full names
        lower_text = text.lower()
        for person in self.people:
            if person.lower() in lower_text and person not in matches:
                matches.append(person)

        # Also match unique first-name aliases inside longer phrases.
        for alias, person in self.alias_to_name.items():
            if " " in alias or len(alias) < 2:
                continue
            if re.search(rf"\b{re.escape(alias)}\b", lower_text) and person not in matches:
                matches.append(person)

        # Handle "OmCar", "BhruguVan" style tokens.
        for vehicle_owner in re.findall(r"\b([A-Za-z]+)(?:car|van)\b", text, flags=re.IGNORECASE):
            person = self._resolve_name(vehicle_owner)
            if person and person not in matches:
                matches.append(person)
        return matches

    def _resolve_name(self, raw: str) -> str | None:
        quick = re.sub(r"[^A-Za-z]", "", raw).lower()
        if quick:
            if quick in self.nickname_aliases:
                return self.nickname_aliases[quick]
            if quick in self.alias_to_name:
                return self.alias_to_name[quick]

        raw = self._normalize_person_name(raw)
        if not raw:
            return None

        if raw.lower() in self.nickname_aliases:
            return self.nickname_aliases[raw.lower()]
        if raw.lower() in self.alias_to_name:
            return self.alias_to_name[raw.lower()]

        first = raw.split()[0].lower()
        if first in self.alias_to_name:
            return self.alias_to_name[first]

        for person in self.people:
            if raw.lower() in person.lower():
                return person
        return None

    def _normalize_person_name(self, raw: str) -> str:
        cleaned = re.sub(r"\s+", " ", raw).strip()
        cleaned = re.sub(r"[^A-Za-z\-\s]", "", cleaned)
        if not cleaned:
            return ""
        if cleaned.lower() in {"name", "team", "captain name", "liaison name", "ras judges"}:
            return ""
        if len(cleaned) < 2:
            return ""
        if len(cleaned.split()) > 4:
            return ""
        if len(cleaned.split()) == 1 and len(cleaned) < 4:
            return ""
        return " ".join(part.capitalize() for part in cleaned.split())

    def _register_person(
        self,
        raw_name: str,
        category: str = "other",
        committee: str = "",
        role: str = "",
        phone: str = "",
        section: str = "",
        board: bool = False,
        liaison: bool = False,
        volunteer: bool = False,
    ) -> None:
        person = self._normalize_person_name(raw_name)
        if not person:
            return

        self.people.add(person)
        self.alias_to_name.setdefault(person.lower(), person)

        first = person.split()[0].lower()
        if first not in self.alias_to_name:
            self.alias_to_name[first] = person
        elif self.alias_to_name[first] != person:
            self.alias_to_name[first] = ""

        if board:
            self.board_members.add(person)
            self.person_category[person] = "board"
        elif liaison:
            self.person_category.setdefault(person, "liaison")
        elif volunteer:
            self.person_category.setdefault(person, "volunteer")
            self.volunteers.add(person)
        else:
            self.person_category.setdefault(person, category)
        if liaison:
            self.liaisons.add(person)
            if person not in self.board_members:
                self.person_category[person] = "liaison"
        if committee and board:
            self.person_committee[person] = committee.strip()

        if not any(c["name"] == person and c["phone"] == phone and c["role"] == role for c in self.contacts):
            self.contacts.append(
                {
                    "name": person,
                    "category": self.person_category.get(person, category),
                    "committee": self.person_committee.get(person, committee).strip(),
                    "role": role.strip(),
                    "phone": phone.strip(),
                    "section": section.strip(),
                }
            )

    def _add_task(self, assignee: str, task: Task) -> None:
        # Basic dedupe key keeps outputs cleaner.
        existing = self.tasks_by_assignee[assignee]
        key = (task.sheet, task.when, task.summary, task.role)
        for t in existing:
            if (t.sheet, t.when, t.summary, t.role) == key:
                return
        existing.append(task)

    def _sorted_tasks(self, tasks: Iterable[Task]) -> list[Task]:
        return sorted(tasks, key=self._task_sort_key)

    def _task_sort_key(self, task: Task) -> tuple:
        day_index = len(DAY_ORDER) + 1
        for day_name, idx in DAY_ORDER.items():
            if day_name in task.sheet:
                day_index = idx
                break
        return (day_index, self._time_sort_value(task.when), task.sheet, task.role, task.summary)

    def _time_sort_value(self, value: str) -> int:
        if not value:
            return 24 * 60 + 1
        test = value.strip().replace("[", "").replace("]", "")
        match = re.search(r"(\d{1,2}:\d{2}\s*[APMapm]{2})", test)
        if not match:
            return 24 * 60 + 2
        try:
            parsed = datetime.strptime(match.group(1).upper().replace(" ", ""), "%I:%M%p")
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            return 24 * 60 + 3

    def _to_export(self) -> dict[str, object]:
        people_export = {}
        for person in sorted(self.people):
            tasks = self._sorted_tasks(self.tasks_by_assignee.get(person, []))
            people_export[person] = {
                "group_membership": {
                    "board": person in self.board_members,
                    "liaison": person in self.liaisons,
                    "volunteer": person in self.volunteers,
                },
                "category": self.person_category.get(person, "other"),
                "committee": self.person_committee.get(person, ""),
                "tasks": [t.to_dict() for t in tasks],
            }

        shared = {
            "board_tasks": [t.to_dict() for t in self._sorted_tasks(self.tasks_by_assignee.get(GROUP_BOARD, []))],
            "liaison_tasks": [t.to_dict() for t in self._sorted_tasks(self.tasks_by_assignee.get(GROUP_LIAISON, []))],
            "volunteer_tasks": [t.to_dict() for t in self._sorted_tasks(self.tasks_by_assignee.get(GROUP_VOLUNTEER, []))],
        }

        return {
            "source_file": str(self.workbook_path),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "important_information": self.important_info,
            "contacts": sorted(self.contacts, key=lambda c: (c["category"], c["name"])),
            "groups": {
                "board_members": sorted(self.board_members),
                "liaisons": sorted(self.liaisons),
                "volunteers": sorted(self.volunteers),
            },
            "shared_tasks": shared,
            "calendar": self.calendar_by_sheet,
            "people": people_export,
        }


def write_outputs(data: dict[str, object], outdir: Path) -> None:
    outdir.mkdir(parents=True, exist_ok=True)
    todos_dir = outdir / "todos"
    todos_dir.mkdir(exist_ok=True)
    for old_file in todos_dir.glob("*.md"):
        old_file.unlink()

    with (outdir / "all_todos.json").open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    important = data.get("important_information", [])
    people = data.get("people", {})

    for person, payload in people.items():
        tasks = payload.get("tasks", [])
        membership = payload.get("group_membership", {})
        category = payload.get("category", "other")
        committee = payload.get("committee", "")
        lines = [f"# {person} - Personal To-Do", ""]
        lines.append("## Important Information")
        if important:
            for item in important:
                bits = [item.get("item", ""), item.get("location", ""), item.get("address", "")]
                bits = [b for b in bits if b]
                lines.append(f"- {' | '.join(bits)}")
        else:
            lines.append("- No important information found")

        lines.append("")
        lines.append("## Responsibilities")
        if tasks:
            for task in tasks:
                when = f"[{task.get('when')}] " if task.get("when") else ""
                details = f" ({task.get('details')})" if task.get("details") else ""
                lines.append(
                    f"- {when}{task.get('summary')} | {task.get('sheet')} | column: {task.get('role')}{details}"
                )
        else:
            lines.append("- No responsibilities detected from workbook")

        lines.append("")
        lines.append("## Group Membership")
        lines.append(f"- Category: {category}")
        if committee:
            lines.append(f"- Committee: {committee}")
        lines.append(f"- Board: {'Yes' if membership.get('board') else 'No'}")
        lines.append(f"- Liaison: {'Yes' if membership.get('liaison') else 'No'}")

        slug = slugify(person)
        (todos_dir / f"{slug}.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    (outdir / "index.html").write_text(build_calendar_viewer(data), encoding="utf-8")
    (outdir / "calendar.html").write_text(build_calendar_viewer(data), encoding="utf-8")
    (outdir / "personal.html").write_text(build_html_viewer(data), encoding="utf-8")
    (outdir / "info.html").write_text(build_info_viewer(data), encoding="utf-8")


def build_html_viewer(data: dict[str, object]) -> str:
    payload = json.dumps(data).replace("</", "<\\/")
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>🎯 Personal Gitas</title>
  <style>
    :root {{
      --bg: #f6f2fb;
      --card: #ffffff;
      --text: #1b2530;
      --muted: #5e6f82;
      --accent: #0f78d8;
      --accent-2: #1a9c74;
      --border: #dde3ef;
      --time-bg: #e8f8f2;
      --time-text: #0b6e4f;
    }}
    body {{
      margin: 0;
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      background: linear-gradient(180deg, #eef7ff, var(--bg));
      color: var(--text);
    }}
    .wrap {{
      max-width: 1200px;
      margin: 0 auto;
      padding: 24px 16px 40px;
    }}
    .card {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 14px;
      padding: 16px;
      margin-bottom: 14px;
      box-shadow: 0 6px 20px rgba(10, 20, 30, 0.06);
    }}
    h1 {{ margin: 0 0 12px; font-size: 1.7rem; }}
    h2 {{ margin: 8px 0 10px; font-size: 1.1rem; }}
    .row {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
    select, input {{
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 10px;
      font-size: 0.95rem;
      min-width: 220px;
    }}
    .meta {{ color: var(--muted); font-size: 0.88rem; margin-top: 8px; }}
    ul {{ margin: 8px 0 0 18px; padding: 0; }}
    li {{ margin: 6px 0; }}
    .tag {{
      display: inline-block;
      margin-right: 8px;
      padding: 3px 8px;
      border-radius: 999px;
      background: #e7f6ef;
      color: var(--accent);
      font-size: 0.8rem;
      border: 1px solid #c0e8d8;
    }}
    .tabs {{ display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 12px; }}
    .tab {{
      text-decoration: none;
      color: var(--text);
      background: #f7f9fc;
      border: 1px solid var(--border);
      border-radius: 999px;
      padding: 7px 12px;
      font-weight: 600;
    }}
    .tab.active {{ background: #e8f4ff; color: var(--accent); border-color: #c8def5; }}
    .task {{
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 12px;
      margin-bottom: 10px;
      background: #fff;
    }}
    .task-head {{
      display: flex;
      gap: 8px;
      align-items: center;
      flex-wrap: wrap;
      margin-bottom: 6px;
    }}
    .time-tag {{
      background: var(--time-bg);
      color: var(--time-text);
      border: 1px solid #b6e2cf;
      border-radius: 999px;
      padding: 3px 10px;
      font-size: 0.8rem;
      font-weight: 600;
    }}
    .day-tag {{
      background: #eef4ff;
      color: #285f9f;
      border: 1px solid #cfe0f9;
      border-radius: 999px;
      padding: 3px 10px;
      font-size: 0.8rem;
      font-weight: 600;
    }}
    .title {{ font-weight: 700; }}
    .subtle {{ color: var(--muted); font-size: 0.86rem; }}
    .kv {{ margin-top: 6px; font-size: 0.92rem; }}
    .kv b {{ color: #2e3d48; }}
    .layout-buttons {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin-top: 8px;
    }}
    .layout-buttons button {{
      border: 1px solid var(--border);
      border-radius: 999px;
      background: #f8fbff;
      color: var(--text);
      padding: 7px 11px;
      cursor: pointer;
    }}
    .layout-buttons button.active {{ background: #e8f4ff; color: var(--accent); border-color: #c8def5; }}
    .grid {{
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    }}
    .day-block {{ margin-bottom: 10px; }}
    .day-title {{
      font-weight: 700;
      margin: 8px 0;
      color: var(--accent-2);
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="tabs">
      <a class="tab" href="./index.html">🗓️ Calendar</a>
      <a class="tab active" href="./personal.html">🎯 Personal Gitas</a>
      <a class="tab" href="./info.html">📇 Contacts & Info</a>
    </div>
    <div class="card">
      <h1>🎯 Personal Gitas</h1>
      <div class="row">
        <select id="category"></select>
        <select id="committee"></select>
        <input id="nameSearch" placeholder="Search your name..." />
        <select id="person"></select>
        <input id="taskSearch" placeholder="Filter tasks..." />
      </div>
      <div id="summary" class="meta"></div>
      <div class="layout-buttons">
        <button id="layoutList" class="active">Vertical List</button>
        <button id="layoutGrid">Horizontal Cards</button>
        <button id="layoutDay">Group By Day</button>
      </div>
    </div>
    <div class="card">
      <h2>📍 Important Information</h2>
      <ul id="important"></ul>
    </div>
    <div class="card">
      <h2>✅ Responsibilities</h2>
      <div id="tasks"></div>
    </div>
  </div>
  <script>
    const data = {payload};
    const categories = ["all", "board", "liaison", "volunteer", "other"];
    const committees = [...new Set(Object.values(data.people).map(p => p.committee).filter(Boolean))].sort();
    const people = Object.keys(data.people).sort((a, b) => a.localeCompare(b));
    const categoryEl = document.getElementById("category");
    const committeeEl = document.getElementById("committee");
    const nameSearch = document.getElementById("nameSearch");
    const select = document.getElementById("person");
    const taskSearch = document.getElementById("taskSearch");
    const layoutList = document.getElementById("layoutList");
    const layoutGrid = document.getElementById("layoutGrid");
    const layoutDay = document.getElementById("layoutDay");
    const summary = document.getElementById("summary");
    const importantEl = document.getElementById("important");
    const tasksEl = document.getElementById("tasks");
    let currentLayout = "list";

    function renderImportant() {{
      importantEl.innerHTML = "";
      data.important_information.forEach(item => {{
        const li = document.createElement("li");
        li.textContent = [item.item, item.location, item.address].filter(Boolean).join(" | ");
        importantEl.appendChild(li);
      }});
    }}

    function formatTask(task) {{
      const parsed = task.parsed || {{}};
      const format = parsed.format || "raw";

      const taskDiv = document.createElement("div");
      taskDiv.className = "task";

      const head = document.createElement("div");
      head.className = "task-head";
      const dayTag = document.createElement("span");
      dayTag.className = "day-tag";
      dayTag.textContent = (task.sheet || "").replace(" Logistics", "");
      const timeTag = document.createElement("span");
      timeTag.className = "time-tag";
      timeTag.textContent = task.when || parsed.time || "No Time";
      const title = document.createElement("span");
      title.className = "title";
      title.textContent = parsed.title || task.summary.split("\\n")[0];
      head.appendChild(dayTag);
      head.appendChild(timeTag);
      head.appendChild(title);
      taskDiv.appendChild(head);

      if (format === "route_depart") {{
        const body = document.createElement("div");
        body.className = "kv";
        body.innerHTML = `<b>People:</b> ${{parsed.people || "-"}}<br><b>Vehicle:</b> ${{parsed.car || "-"}}<br><b>Route:</b> ${{parsed.from || "-"}} -> ${{parsed.to || "-"}}${{parsed.purpose ? `<br><b>Purpose:</b> ${{parsed.purpose}}` : ""}}`;
        taskDiv.appendChild(body);
      }} else if (format === "route_arrive") {{
        const body = document.createElement("div");
        body.className = "kv";
        body.innerHTML = `<b>People:</b> ${{parsed.people || "-"}}<br><b>Vehicle:</b> ${{parsed.car || "-"}}<br><b>Arrival:</b> ${{parsed.at || "-"}}${{parsed.purpose ? `<br><b>Purpose:</b> ${{parsed.purpose}}` : ""}}`;
        taskDiv.appendChild(body);
      }} else {{
        const raw = document.createElement("div");
        raw.className = "kv";
        raw.style.whiteSpace = "pre-wrap";
        raw.textContent = parsed.body || parsed.raw || task.summary;
        taskDiv.appendChild(raw);
      }}

      const subtle = document.createElement("div");
      subtle.className = "subtle";
      subtle.textContent = `${{task.sheet}} | column: ${{task.role}}${{task.details ? ` | ${{task.details}}` : ""}}`;
      taskDiv.appendChild(subtle);

      const full = parsed.full_text || task.summary;
      if (full) {{
        const details = document.createElement("details");
        const summaryEl = document.createElement("summary");
        summaryEl.textContent = "Full cell text";
        const pre = document.createElement("div");
        pre.className = "kv";
        pre.style.whiteSpace = "pre-wrap";
        pre.textContent = full;
        details.appendChild(summaryEl);
        details.appendChild(pre);
        taskDiv.appendChild(details);
      }}
      return taskDiv;
    }}

    function refreshPersonOptions() {{
      const cat = categoryEl.value;
      const comm = committeeEl.value;
      const needle = nameSearch.value.trim().toLowerCase();
      const current = select.value;
      select.innerHTML = "";

      const filtered = people.filter(name => {{
        const p = data.people[name];
        const categoryMatch = cat === "all" || p.category === cat;
        const committeeMatch = comm === "All Committees" || !comm || p.committee === comm;
        const nameMatch = !needle || name.toLowerCase().includes(needle);
        return categoryMatch && committeeMatch && nameMatch;
      }});

      filtered.forEach(name => {{
        const opt = document.createElement("option");
        const p = data.people[name];
        const committeeLabel = p.committee ? ` (${{p.committee}})` : "";
        const categoryLabel = p.category ? p.category.charAt(0).toUpperCase() + p.category.slice(1) : "Other";
        opt.value = name;
        opt.textContent = `${{name}} [${{categoryLabel}}]${{committeeLabel}}`;
        select.appendChild(opt);
      }});

      if (filtered.includes(current)) {{
        select.value = current;
      }} else if (filtered.length) {{
        select.value = filtered[0];
      }}
    }}

    categories.forEach(cat => {{
      const opt = document.createElement("option");
      opt.value = cat;
      opt.textContent = cat.charAt(0).toUpperCase() + cat.slice(1);
      categoryEl.appendChild(opt);
    }});
    ["All Committees", ...committees].forEach(comm => {{
      const opt = document.createElement("option");
      opt.value = comm;
      opt.textContent = comm;
      committeeEl.appendChild(opt);
    }});
    categoryEl.value = "all";
    committeeEl.value = "All Committees";
    refreshPersonOptions();
    renderImportant();
    renderPerson();

    categoryEl.addEventListener("change", () => {{ refreshPersonOptions(); renderPerson(); }});
    committeeEl.addEventListener("change", () => {{ refreshPersonOptions(); renderPerson(); }});
    nameSearch.addEventListener("input", () => {{ refreshPersonOptions(); renderPerson(); }});
    select.addEventListener("change", renderPerson);
    taskSearch.addEventListener("input", renderPerson);
    function setLayout(next) {{
      currentLayout = next;
      [layoutList, layoutGrid, layoutDay].forEach(btn => btn.classList.remove("active"));
      if (next === "list") layoutList.classList.add("active");
      if (next === "grid") layoutGrid.classList.add("active");
      if (next === "day") layoutDay.classList.add("active");
      renderPerson();
    }}
    layoutList.addEventListener("click", () => setLayout("list"));
    layoutGrid.addEventListener("click", () => setLayout("grid"));
    layoutDay.addEventListener("click", () => setLayout("day"));
    function renderPerson() {{
      const person = select.value;
      if (!person || !data.people[person]) {{
        tasksEl.innerHTML = "<div class='subtle'>No people match the current filters.</div>";
        summary.textContent = "0 tasks";
        return;
      }}
      const payload = data.people[person];
      const needle = taskSearch.value.trim().toLowerCase();
      tasksEl.innerHTML = "";
      const tags = [];
      if (payload.group_membership.board) tags.push('<span class="tag">Board</span>');
      if (payload.group_membership.liaison) tags.push('<span class="tag">Liaison</span>');
      if (payload.category === "volunteer") tags.push('<span class="tag">Volunteer</span>');
      if (payload.committee) tags.push(`<span class="tag">${{payload.committee}}</span>`);
      summary.innerHTML = `${{tags.join("")}}${{payload.tasks.length}} tasks`;
      const filtered = payload.tasks.filter(t => !needle || `${{t.when}} ${{t.summary}} ${{t.details}} ${{t.role}} ${{t.sheet}}`.toLowerCase().includes(needle));
      if (!filtered.length) {{
        tasksEl.innerHTML = "<div class='subtle'>No tasks match this filter.</div>";
        return;
      }}
      if (currentLayout === "grid") {{
        const wrap = document.createElement("div");
        wrap.className = "grid";
        filtered.forEach(task => wrap.appendChild(formatTask(task)));
        tasksEl.appendChild(wrap);
      }} else if (currentLayout === "day") {{
        const byDay = {{}};
        filtered.forEach(t => {{
          if (!byDay[t.sheet]) byDay[t.sheet] = [];
          byDay[t.sheet].push(t);
        }});
        Object.keys(byDay).forEach(day => {{
          const block = document.createElement("div");
          block.className = "day-block";
          const h = document.createElement("div");
          h.className = "day-title";
          h.textContent = day;
          block.appendChild(h);
          byDay[day].forEach(task => block.appendChild(formatTask(task)));
          tasksEl.appendChild(block);
        }});
      }} else {{
        filtered.forEach(task => tasksEl.appendChild(formatTask(task)));
      }}
    }}
    renderPerson();
  </script>
</body>
</html>
"""


def build_calendar_viewer(data: dict[str, object]) -> str:
    payload = json.dumps(data.get("calendar", {})).replace("</", "<\\/")
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>🗓️ Masti Calendar</title>
  <style>
    :root {{
      --bg: #f2f6fb;
      --card: #ffffff;
      --text: #1e2a35;
      --muted: #657786;
      --accent: #1c6ca1;
      --border: #d2dbe6;
    }}
    body {{
      margin: 0;
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      background: linear-gradient(180deg, #eaf4ff, var(--bg));
      color: var(--text);
    }}
    .wrap {{
      max-width: 1200px;
      margin: 0 auto;
      padding: 14px 10px 28px;
    }}
    .card {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 14px;
      padding: 16px;
      margin-bottom: 14px;
    }}
    .tabs {{ display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 12px; }}
    .tab {{
      text-decoration: none;
      color: var(--text);
      background: #f7f9fc;
      border: 1px solid var(--border);
      border-radius: 999px;
      padding: 7px 12px;
      font-weight: 600;
    }}
    .tab.active {{ background: #e8f4ff; color: var(--accent); border-color: #c8def5; }}
    .row {{ display: flex; gap: 10px; flex-wrap: wrap; }}
    input {{
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 9px;
      font-size: 0.92rem;
      min-width: 230px;
    }}
    .day-title {{ margin: 0 0 10px; color: var(--accent); font-size: 1.12rem; }}
    .muted {{ color: var(--muted); }}
    .table-wrap {{
      width: 100%;
      overflow-x: auto;
      -webkit-overflow-scrolling: touch;
    }}
    table {{
      border-collapse: collapse;
      width: max-content;
      min-width: 100%;
      table-layout: fixed;
      font-size: 0.84rem;
      background: #fff;
    }}
    th, td {{
      border: 1px solid var(--border);
      padding: 6px;
      vertical-align: top;
      white-space: pre-wrap;
      line-height: 1.25;
    }}
    th {{
      background: #f4f8fc;
      font-weight: 700;
      color: #264d73;
      position: sticky;
      top: 0;
      z-index: 1;
    }}
    td.time-col, th.time-col {{
      width: 78px;
      min-width: 78px;
      max-width: 78px;
      text-align: center;
      font-weight: 700;
      background: #f8fbff;
    }}
    td.data-col, th.data-col {{
      width: 160px;
      min-width: 160px;
    }}
    @media (max-width: 600px) {{
      table {{ font-size: 0.8rem; }}
      td.data-col, th.data-col {{ width: 140px; min-width: 140px; }}
      td.time-col, th.time-col {{ width: 68px; min-width: 68px; max-width: 68px; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="tabs">
      <a class="tab active" href="./index.html">🗓️ Calendar</a>
      <a class="tab" href="./personal.html">🎯 Personal Gitas</a>
      <a class="tab" href="./info.html">📇 Contacts & Info</a>
    </div>
    <div class="card">
      <h1 style="margin:0 0 10px;">🗓️ Full Logistics Calendar (Wed-Sun)</h1>
      <div class="row">
        <input id="q" placeholder="Search name or keyword across all days..." />
      </div>
      <div class="muted" id="hits"></div>
    </div>
    <div id="calendar"></div>
  </div>
  <script>
    const calendar = {payload};
    const container = document.getElementById("calendar");
    const q = document.getElementById("q");
    const hits = document.getElementById("hits");

    function render() {{
      const needle = q.value.trim().toLowerCase();
      container.innerHTML = "";
      let total = 0;
      Object.values(calendar).forEach(day => {{
        const dayCard = document.createElement("div");
        dayCard.className = "card";
        const h = document.createElement("h2");
        h.className = "day-title";
        h.textContent = day.title || day.sheet;
        dayCard.appendChild(h);

        const wrap = document.createElement("div");
        wrap.className = "table-wrap";
        const table = document.createElement("table");
        const thead = document.createElement("thead");
        const hr = document.createElement("tr");
        const timeHead = document.createElement("th");
        timeHead.className = "time-col";
        timeHead.textContent = "TIME";
        hr.appendChild(timeHead);
        const cols = (day.headers || []).filter(h => h !== "TIME");
        cols.forEach(col => {{
          const th = document.createElement("th");
          th.className = "data-col";
          th.textContent = col || "Unlabeled";
          hr.appendChild(th);
        }});
        thead.appendChild(hr);
        table.appendChild(thead);
        const tbody = document.createElement("tbody");
        (day.events || []).forEach(ev => {{
          const rowText = `${{ev.time}} ${{(ev.cells || []).map(c => c.column + " " + c.text).join(" ")}}`.toLowerCase();
          if (needle && !rowText.includes(needle)) return;
          total += 1;
          const tr = document.createElement("tr");
          const tcell = document.createElement("td");
          tcell.className = "time-col";
          tcell.textContent = ev.time || "";
          tr.appendChild(tcell);
          cols.forEach(col => {{
            const td = document.createElement("td");
            td.className = "data-col";
            const match = (ev.cells || []).find(c => c.column === col);
            td.textContent = match ? match.text : "";
            tr.appendChild(td);
          }});
          tbody.appendChild(tr);
        }});
        table.appendChild(tbody);
        wrap.appendChild(table);
        dayCard.appendChild(wrap);
        container.appendChild(dayCard);
      }});
      hits.textContent = `${{total}} matching calendar rows`;
    }}

    render();
    q.addEventListener("input", render);
  </script>
</body>
</html>
"""


def build_info_viewer(data: dict[str, object]) -> str:
    payload = json.dumps(data).replace("</", "<\\/")
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>📇 Contacts & Info</title>
  <style>
    :root {{
      --bg: #fff8ef;
      --card: #fff;
      --text: #22313c;
      --muted: #61717e;
      --accent: #b35b0f;
      --border: #e7d4c2;
    }}
    body {{
      margin: 0;
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      background: linear-gradient(180deg, #fff3e4, var(--bg));
      color: var(--text);
    }}
    .wrap {{ max-width: 1200px; margin: 0 auto; padding: 24px 16px 40px; }}
    .tabs {{ display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 12px; }}
    .tab {{
      text-decoration: none;
      color: var(--text);
      background: #fff8f0;
      border: 1px solid var(--border);
      border-radius: 999px;
      padding: 7px 12px;
      font-weight: 600;
    }}
    .tab.active {{ background: #fff1dd; color: var(--accent); }}
    .card {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 14px;
      padding: 16px;
      margin-bottom: 14px;
    }}
    table {{ width: 100%; border-collapse: collapse; font-size: 0.92rem; }}
    th, td {{ border: 1px solid var(--border); padding: 8px; text-align: left; }}
    th {{ background: #fff4e6; position: sticky; top: 0; }}
    .muted {{ color: var(--muted); }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="tabs">
      <a class="tab" href="./index.html">🗓️ Calendar</a>
      <a class="tab" href="./personal.html">🎯 Personal Gitas</a>
      <a class="tab active" href="./info.html">📇 Contacts & Info</a>
    </div>
    <div class="card">
      <h1 style="margin:0 0 10px;">📇 Contacts</h1>
      <div class="muted">Board, liaisons, volunteers, and other contacts from the workbook.</div>
      <table id="contactsTable">
        <thead>
          <tr><th>Name</th><th>Category</th><th>Committee</th><th>Role</th><th>Phone</th><th>Section</th></tr>
        </thead>
        <tbody></tbody>
      </table>
    </div>
    <div class="card">
      <h1 style="margin:0 0 10px;">📍 Important Information</h1>
      <table id="infoTable">
        <thead><tr><th>Item</th><th>Location</th><th>Address</th></tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>
  <script>
    const data = {payload};
    const cbody = document.querySelector("#contactsTable tbody");
    const ibody = document.querySelector("#infoTable tbody");
    (data.contacts || []).forEach(c => {{
      const tr = document.createElement("tr");
      ["name","category","committee","role","phone","section"].forEach(key => {{
        const td = document.createElement("td");
        let value = c[key] || "";
        if (key === "category" && value) value = value.charAt(0).toUpperCase() + value.slice(1);
        td.textContent = value;
        tr.appendChild(td);
      }});
      cbody.appendChild(tr);
    }});
    (data.important_information || []).forEach(i => {{
      const tr = document.createElement("tr");
      ["item","location","address"].forEach(key => {{
        const td = document.createElement("td");
        td.textContent = i[key] || "";
        tr.appendChild(td);
      }});
      ibody.appendChild(tr);
    }});
  </script>
</body>
</html>
"""


def main() -> None:
    parser = argparse.ArgumentParser(description="Build personal to-do lists from competition workbook")
    parser.add_argument("xlsx", type=Path, help="Path to source .xlsx file")
    parser.add_argument("--out", type=Path, default=Path("output"), help="Output directory")
    args = parser.parse_args()

    builder = TodoBuilder(args.xlsx)
    data = builder.build()
    write_outputs(data, args.out)

    people_count = len(data.get("people", {}))
    board_tasks = len(data.get("shared_tasks", {}).get("board_tasks", []))
    liaison_tasks = len(data.get("shared_tasks", {}).get("liaison_tasks", []))
    print(f"Generated to-dos for {people_count} people")
    print(f"Shared board tasks: {board_tasks}")
    print(f"Shared liaison tasks: {liaison_tasks}")
    print(f"Output directory: {args.out}")


if __name__ == "__main__":
    main()
