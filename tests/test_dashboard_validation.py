from __future__ import annotations

import os
import re
import sys
import unittest
from pathlib import Path

import openpyxl

from gita_reader.pipeline import (
    LOGISTICS_SHEETS,
    TodoBuilder,
)


def _norm(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _find_nearest_time(ws, row_idx: int) -> str:
    for idx in range(row_idx, max(1, row_idx - 6), -1):
        val = _norm(ws.cell(idx, 1).value)
        if val:
            return val
    return ""


def _normalize_text(text: str) -> str:
    return "\n".join(line.rstrip() for line in text.replace("\r\n", "\n").strip().split("\n"))


def _is_all_board(role: str, text: str) -> bool:
    probe = f"{role} {text}".upper()
    return "ALL_BOARD" in probe or "ALL BOARD" in probe


def _is_all_liaison(role: str, text: str) -> bool:
    probe = f"{role} {text}".upper()
    return "ALL_LIAISON" in probe or "ALL LIAISON" in probe or "ALL LIAISONS" in probe


def _is_all_volunteer(role: str, text: str) -> bool:
    probe = f"{role} {text}".upper()
    return "ALL_VOLUNTEER" in probe or "ALL VOLUNTEER" in probe or "ALL VOLUNTEERS" in probe


def _is_directors_log(role: str, text: str) -> bool:
    probe = f"{role} {text}".upper()
    return "DIRECTORS_LOG" in probe or "DIRECTORS LOG" in probe


class DashboardSpreadsheetValidationTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.repo_root = Path(__file__).resolve().parents[1]
        cls.xlsx_path = cls.repo_root / "MM19 Gita.xlsx"
        if not cls.xlsx_path.exists():
            raise unittest.SkipTest(f"Workbook not found at {cls.xlsx_path}")

        cls.builder = TodoBuilder(cls.xlsx_path)
        cls.data = cls.builder.build()
        cls.people = sorted(cls.data["people"].keys())

        groups = cls.data.get("groups", {})
        cls.board_members = set(groups.get("board_members", []))
        cls.liaisons = set(groups.get("liaisons", []))
        cls.volunteers = set(groups.get("volunteers", []))
        cls.directors_members = set(cls.builder.directors_members)
        cls.verbose_logs = os.getenv("GITA_TEST_VERBOSE", "").lower() in {"1", "true", "yes", "on"}
        if cls.verbose_logs:
            cls._log(f"Loaded workbook: {cls.xlsx_path}")
            cls._log(f"People discovered: {len(cls.people)}")

    @classmethod
    def _log(cls, message: str) -> None:
        if cls.verbose_logs:
            print(f"[gita-test] {message}", file=sys.stderr)

    def _expected_logistics_cells_by_person(self) -> dict[str, set[tuple[str, str, str]]]:
        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        expected: dict[str, set[tuple[str, str, str]]] = {person: set() for person in self.people}

        for sheet_name in LOGISTICS_SHEETS:
            ws = wb[sheet_name]
            header_row = None
            headers: dict[int, str] = {}

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
                values = [_norm(v) for v in row]
                if "TIME" in values:
                    header_row = row_idx
                    headers = {i: values[i].strip().upper() for i in range(len(values)) if values[i]}
                    continue
                if header_row is None or row_idx <= header_row:
                    continue
                if not any(values):
                    continue

                when = values[0] or _find_nearest_time(ws, row_idx)
                _ = when  # Kept for future debugging if needed.

                for col_idx, text in enumerate(values[1:], start=1):
                    if not text:
                        continue
                    role = headers.get(col_idx, "General")
                    cell_text = _normalize_text(text)
                    signature = (sheet_name, role, cell_text)

                    explicit = set(self.builder._extract_named_people(cell_text))
                    for person in explicit:
                        expected[person].add(signature)

                    if _is_all_board(role, cell_text):
                        for person in self.board_members:
                            expected[person].add(signature)
                    if _is_all_liaison(role, cell_text):
                        for person in self.liaisons:
                            expected[person].add(signature)
                    if _is_all_volunteer(role, cell_text):
                        for person in self.volunteers:
                            expected[person].add(signature)
                    if _is_directors_log(role, cell_text):
                        for person in self.directors_members:
                            expected[person].add(signature)

        return expected

    def _actual_logistics_cells_by_person(self) -> dict[str, set[tuple[str, str, str]]]:
        actual: dict[str, set[tuple[str, str, str]]] = {person: set() for person in self.people}

        for person, payload in self.data["people"].items():
            for task in payload.get("tasks", []):
                if task.get("sheet") not in LOGISTICS_SHEETS:
                    continue
                parsed = task.get("parsed") or {}
                full_text = parsed.get("full_text") or task.get("summary") or ""
                signature = (
                    task.get("sheet", ""),
                    task.get("role", ""),
                    _normalize_text(full_text),
                )
                actual[person].add(signature)

        return actual

    def test_person_name_search_can_find_every_person(self) -> None:
        # Mirrors the dashboard behavior: lowercased substring match in the name search field.
        people = sorted(self.data["people"].keys())
        self._log("Running name-search validation")
        for person in people:
            self._log(f"name-search person: {person}")
            query = person.split()[0][:3].lower() if person.split() else person.lower()
            filtered = [name for name in people if query in name.lower()]
            self.assertIn(person, filtered, f"Name search failed for {person} with query '{query}'")

    def test_per_person_logistics_cell_set_matches_spreadsheet(self) -> None:
        self._log("Building expected/actual logistics cell maps")
        expected = self._expected_logistics_cells_by_person()
        actual = self._actual_logistics_cells_by_person()

        for person in self.people:
            self._log(
                f"logistics person: {person} | expected={len(expected[person])} actual={len(actual[person])}"
            )
            self.assertSetEqual(
                actual[person],
                expected[person],
                msg=f"Logistics cell mismatch for {person}",
            )

    def test_all_x_cells_are_fully_included_for_each_person(self) -> None:
        self._log("Running ALL_* propagation validation")
        expected = self._expected_logistics_cells_by_person()
        actual = self._actual_logistics_cells_by_person()

        def is_all_x(sig: tuple[str, str, str]) -> bool:
            return "ALL_" in sig[2].upper() or "ALL " in sig[2].upper()

        for person in self.people:
            expected_all = {sig for sig in expected[person] if is_all_x(sig)}
            actual_all = {sig for sig in actual[person] if is_all_x(sig)}
            self._log(
                f"all-x person: {person} | expected_all={len(expected_all)} actual_all={len(actual_all)}"
            )
            self.assertSetEqual(
                actual_all,
                expected_all,
                msg=f"ALL_* cell mismatch for {person}",
            )


if __name__ == "__main__":
    unittest.main()
